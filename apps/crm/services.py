"""
CRM service layer — all business logic lives here.

Views and Celery tasks only call service methods; no ORM in views.
"""
from __future__ import annotations

import datetime
import hashlib
import hmac
import logging
import time
from decimal import Decimal
from typing import Optional
from zoneinfo import ZoneInfo

from django.db import transaction
from django.utils import timezone

logger = logging.getLogger(__name__)

_MSK = ZoneInfo("Europe/Moscow")


# ─── Telegram auth ────────────────────────────────────────────────────────────

class TelegramAuthError(Exception):
    pass


def verify_telegram_login(data: dict, bot_token: str, max_age_seconds: int = 86400) -> dict:
    """
    Verify the Telegram Login Widget payload.

    Raises TelegramAuthError if hash is invalid or data is stale.
    Returns the verified data dict (with 'hash' removed).

    Telegram's verification algorithm:
      1. Remove 'hash' from data
      2. Sort remaining fields, join as "key=value\n..."
      3. SHA256 of bot_token → used as HMAC key
      4. HMAC-SHA256(key, data_check_string) must equal hash
    """
    data = dict(data)
    received_hash = data.pop("hash", "")

    if not received_hash:
        raise TelegramAuthError("Hash missing from Telegram auth data")

    data_check_string = "\n".join(
        f"{k}={v}" for k, v in sorted(data.items())
    )
    secret_key = hashlib.sha256(bot_token.encode()).digest()
    expected_hash = hmac.new(
        secret_key, data_check_string.encode(), hashlib.sha256
    ).hexdigest()

    if not hmac.compare_digest(expected_hash, received_hash):
        raise TelegramAuthError("Telegram auth hash mismatch — possible forgery")

    auth_date = int(data.get("auth_date", 0))
    if time.time() - auth_date > max_age_seconds:
        raise TelegramAuthError("Telegram auth data is stale (> 24h)")

    return data


# ─── Workspace helpers ────────────────────────────────────────────────────────

class WorkspaceService:

    @staticmethod
    def get_or_create_default() -> "Workspace":
        """
        Ensure the default GRAMLY workspace exists.
        Called during first deployment / initial setup.
        """
        from apps.crm.models import Workspace
        ws, _ = Workspace.objects.get_or_create(
            slug="gramly",
            defaults={
                "name": "GRAMLY",
                "description": "Основное рабочее пространство GRAMLY",
            },
        )
        return ws

    @staticmethod
    def get_memberships_for_user(user) -> list:
        from apps.crm.models import WorkspaceMembership
        return list(
            WorkspaceMembership.objects.filter(user=user, is_active=True)
            .select_related("workspace")
            .order_by("workspace__name")
        )

    @staticmethod
    def get_membership(workspace, user) -> Optional["WorkspaceMembership"]:
        from apps.crm.models import WorkspaceMembership
        return WorkspaceMembership.objects.filter(
            workspace=workspace, user=user, is_active=True
        ).first()

    @staticmethod
    def add_member(workspace, user, role: str, invited_by=None) -> "WorkspaceMembership":
        from apps.crm.models import WorkspaceMembership
        membership, created = WorkspaceMembership.objects.get_or_create(
            workspace=workspace,
            user=user,
            defaults={
                "role": role,
                "invited_by": invited_by,
                "joined_at": timezone.now(),
            },
        )
        if not created:
            membership.role = role
            membership.is_active = True
            membership.save(update_fields=["role", "is_active", "updated_at"])
        return membership

    @staticmethod
    def set_member_role(workspace, user, role: str) -> None:
        from apps.crm.models import WorkspaceMembership
        WorkspaceMembership.objects.filter(
            workspace=workspace, user=user
        ).update(role=role)


# ─── Weekly Plan helpers ──────────────────────────────────────────────────────

class WeeklyPlanService:

    @staticmethod
    def get_week_start(date: datetime.date) -> datetime.date:
        """Return the Monday of the ISO week containing `date`."""
        return date - datetime.timedelta(days=date.weekday())

    @staticmethod
    def get_current_plan(workspace) -> Optional["WeeklyPlan"]:
        from apps.crm.models import WeeklyPlan
        today = datetime.datetime.now(tz=_MSK).date()
        week_start = WeeklyPlanService.get_week_start(today)
        return WeeklyPlan.objects.filter(
            workspace=workspace, week_start=week_start
        ).first()

    @staticmethod
    def get_plan_for_date(workspace, date: datetime.date) -> Optional["WeeklyPlan"]:
        from apps.crm.models import WeeklyPlan
        week_start = WeeklyPlanService.get_week_start(date)
        return WeeklyPlan.objects.filter(
            workspace=workspace, week_start=week_start
        ).first()

    @staticmethod
    def upsert_plan(workspace, week_start: datetime.date, pp_plan: Decimal,
                    privat_plan: Decimal, created_by=None) -> "WeeklyPlan":
        from apps.crm.models import WeeklyPlan
        plan, _ = WeeklyPlan.objects.update_or_create(
            workspace=workspace,
            week_start=week_start,
            defaults={
                "pp_plan": pp_plan,
                "privat_plan": privat_plan,
                "created_by": created_by,
            },
        )
        return plan

    @staticmethod
    def get_week_pp_total(workspace, week_start: datetime.date) -> Decimal:
        from apps.crm.models import FinanceEntry
        week_end = week_start + datetime.timedelta(days=6)
        result = FinanceEntry.objects.filter(
            workspace=workspace,
            date__gte=week_start,
            date__lte=week_end,
        ).aggregate(total=models_sum("pp_earnings"))
        return result["total"] or Decimal(0)

    @staticmethod
    def get_week_privat_total(workspace, week_start: datetime.date) -> Decimal:
        from apps.crm.models import FinanceEntry
        week_end = week_start + datetime.timedelta(days=6)
        result = FinanceEntry.objects.filter(
            workspace=workspace,
            date__gte=week_start,
            date__lte=week_end,
        ).aggregate(total=models_sum("privat_earnings"))
        return result["total"] or Decimal(0)


def models_sum(field: str):
    from django.db.models import Sum
    return Sum(field)


# ─── Entry services ───────────────────────────────────────────────────────────

class EntryService:

    @staticmethod
    def get_or_init_finance(workspace, date: datetime.date):
        """Return existing FinanceEntry or an unsaved instance."""
        from apps.crm.models import FinanceEntry
        return FinanceEntry.objects.filter(workspace=workspace, date=date).first()

    @staticmethod
    def get_or_init_application(workspace, date: datetime.date):
        from apps.crm.models import ApplicationEntry
        return ApplicationEntry.objects.filter(workspace=workspace, date=date).first()

    @staticmethod
    @transaction.atomic
    def save_finance_entry(workspace, date: datetime.date, user, data: dict) -> "FinanceEntry":
        from apps.crm.models import FinanceEntry
        entry, created = FinanceEntry.objects.get_or_create(
            workspace=workspace,
            date=date,
            defaults={"submitted_by": user, **data},
        )
        if not created:
            for field, value in data.items():
                setattr(entry, field, value)
            entry.last_edited_at = timezone.now()
            entry.save()
        # Try to generate report if both entries exist
        EntryService._try_generate_report(workspace, date, user)
        return entry

    @staticmethod
    @transaction.atomic
    def save_application_entry(workspace, date: datetime.date, user, data: dict) -> "ApplicationEntry":
        from apps.crm.models import ApplicationEntry
        entry, created = ApplicationEntry.objects.get_or_create(
            workspace=workspace,
            date=date,
            defaults={"submitted_by": user, **data},
        )
        if not created:
            for field, value in data.items():
                setattr(entry, field, value)
            entry.last_edited_at = timezone.now()
            entry.save()
        EntryService._try_generate_report(workspace, date, user)
        return entry

    @staticmethod
    def _try_generate_report(workspace, date: datetime.date, triggered_by=None) -> Optional["DailySummaryReport"]:
        from apps.crm.models import FinanceEntry, ApplicationEntry
        fin = FinanceEntry.objects.filter(workspace=workspace, date=date).first()
        app = ApplicationEntry.objects.filter(workspace=workspace, date=date).first()
        if fin and app:
            return ReportService.generate(workspace, date, fin, app, generated_by=triggered_by)
        return None


# ─── Report service ───────────────────────────────────────────────────────────

class ReportService:

    @staticmethod
    @transaction.atomic
    def generate(workspace, date: datetime.date, finance_entry, application_entry,
                 generated_by=None) -> "DailySummaryReport":
        """
        Build or rebuild DailySummaryReport from the two entry objects.
        Idempotent: updates existing report if called again.
        Also queues Telegram notification to workspace owners.
        """
        from apps.crm.models import DailySummaryReport

        # Compute weekly plan percentages
        plan = WeeklyPlanService.get_plan_for_date(workspace, date)
        week_start = WeeklyPlanService.get_week_start(date)

        pp_week_total     = WeeklyPlanService.get_week_pp_total(workspace, week_start)
        privat_week_total = WeeklyPlanService.get_week_privat_total(workspace, week_start)

        pp_plan     = plan.pp_plan     if plan else Decimal(0)
        privat_plan = plan.privat_plan if plan else Decimal(0)

        pp_pct     = (pp_week_total / pp_plan * 100)     if pp_plan     else Decimal(0)
        privat_pct = (privat_week_total / privat_plan * 100) if privat_plan else Decimal(0)

        balance = finance_entry.income - finance_entry.expenses

        report_text = ReportService._build_text(
            date=date,
            pp_earnings=finance_entry.pp_earnings,
            privat_earnings=finance_entry.privat_earnings,
            pp_pct=pp_pct,
            privat_pct=privat_pct,
            applications_count=application_entry.applications_count,
            applications_earnings=application_entry.applications_earnings,
            income=finance_entry.income,
            expenses=finance_entry.expenses,
            balance=balance,
        )

        is_new = False
        report, created = DailySummaryReport.objects.update_or_create(
            workspace=workspace,
            date=date,
            defaults={
                "finance_entry":          finance_entry,
                "application_entry":      application_entry,
                "pp_earnings":            finance_entry.pp_earnings,
                "privat_earnings":        finance_entry.privat_earnings,
                "pp_plan_pct":            pp_pct.quantize(Decimal("0.01")),
                "privat_plan_pct":        privat_pct.quantize(Decimal("0.01")),
                "applications_count":     application_entry.applications_count,
                "applications_earnings":  application_entry.applications_earnings,
                "cash_flow_income":       finance_entry.income,
                "cash_flow_expenses":     finance_entry.expenses,
                "cash_flow_balance":      balance,
                "report_text":            report_text,
                "generated_by":           generated_by,
            },
        )
        is_new = created

        # Queue Telegram notification for owners (only on fresh generation)
        if is_new:
            from apps.crm.tasks import send_crm_report_notification_task
            send_crm_report_notification_task.delay(report.pk)

        return report

    @staticmethod
    def _build_text(*, date, pp_earnings, privat_earnings, pp_pct, privat_pct,
                    applications_count, applications_earnings, income, expenses, balance) -> str:
        sign = "+" if balance >= 0 else ""
        return (
            f"📊 Ежедневный отчёт • {date.strftime('%d.%m.%Y')}\n\n"
            f"💳 Заработок с ПП (за сегодня): {pp_earnings:.2f} $\n"
            f"🏦 Заработок с Привата (за сегодня): {privat_earnings:.2f} $\n\n"
            f"📈 % выполнения недельного плана ПП: {pp_pct:.1f}%\n"
            f"📈 % выполнения недельного плана Привата: {privat_pct:.1f}%\n\n"
            f"📋 Число заявок за день: {applications_count} шт.\n"
            f"💰 Заработок с заявок за день: {applications_earnings:.2f} $\n\n"
            f"⚖️ Сальдо за день:\n"
            f"   Доход Cash Flow: +{income:.2f} $\n"
            f"   Расходы/выплаты: -{expenses:.2f} $\n"
            f"   Итого: {sign}{balance:.2f} $"
        )

    @staticmethod
    def get_for_workspace(workspace, start_date: datetime.date,
                          end_date: datetime.date) -> list:
        from apps.crm.models import DailySummaryReport
        return list(
            DailySummaryReport.objects.filter(
                workspace=workspace,
                date__gte=start_date,
                date__lte=end_date,
            ).order_by("-date")
        )


# ─── History / dashboard helpers ──────────────────────────────────────────────

class DashboardService:

    @staticmethod
    def get_today_status(workspace) -> dict:
        """Return completion status for today."""
        from apps.crm.models import FinanceEntry, ApplicationEntry, DailySummaryReport
        today = datetime.datetime.now(tz=_MSK).date()
        fin = FinanceEntry.objects.filter(workspace=workspace, date=today).first()
        app = ApplicationEntry.objects.filter(workspace=workspace, date=today).first()
        report = DailySummaryReport.objects.filter(workspace=workspace, date=today).first()
        return {
            "today":           today,
            "finance_entry":   fin,
            "app_entry":       app,
            "finance_done":    fin is not None,
            "apps_done":       app is not None,
            "report":          report,
            "both_done":       fin is not None and app is not None,
        }

    @staticmethod
    def get_recent_reports(workspace, days: int = 7) -> list:
        from apps.crm.models import DailySummaryReport
        cutoff = datetime.datetime.now(tz=_MSK).date() - datetime.timedelta(days=days)
        return list(
            DailySummaryReport.objects.filter(
                workspace=workspace, date__gte=cutoff
            ).order_by("-date")
        )

    @staticmethod
    def get_history_entries(workspace, start: datetime.date, end: datetime.date) -> list:
        """
        Return a merged day-by-day list for the history page.
        Each item: {date, finance, application, report}
        """
        from apps.crm.models import FinanceEntry, ApplicationEntry, DailySummaryReport

        fins  = {e.date: e for e in FinanceEntry.objects.filter(
            workspace=workspace, date__gte=start, date__lte=end
        )}
        apps  = {e.date: e for e in ApplicationEntry.objects.filter(
            workspace=workspace, date__gte=start, date__lte=end
        )}
        reps  = {r.date: r for r in DailySummaryReport.objects.filter(
            workspace=workspace, date__gte=start, date__lte=end
        )}

        days = []
        current = end
        while current >= start:
            days.append({
                "date":        current,
                "finance":     fins.get(current),
                "application": apps.get(current),
                "report":      reps.get(current),
            })
            current -= datetime.timedelta(days=1)
        return days


# ─── Deadline tracking ────────────────────────────────────────────────────────

class DeadlineService:

    @staticmethod
    def check_and_record(workspace) -> Optional["DeadlineMiss"]:
        """
        Check if entries for yesterday (in MSK) are complete.
        If not, create/update a DeadlineMiss record.
        Called by Celery at 00:05 MSK.
        """
        from apps.crm.models import FinanceEntry, ApplicationEntry, DeadlineMiss
        yesterday = (datetime.datetime.now(tz=_MSK) - datetime.timedelta(days=1)).date()

        fin_exists = FinanceEntry.objects.filter(workspace=workspace, date=yesterday).exists()
        app_exists = ApplicationEntry.objects.filter(workspace=workspace, date=yesterday).exists()

        if fin_exists and app_exists:
            return None  # all good

        miss, _ = DeadlineMiss.objects.update_or_create(
            workspace=workspace,
            date=yesterday,
            defaults={
                "finance_missing":      not fin_exists,
                "applications_missing": not app_exists,
                "notified_at":          timezone.now(),
            },
        )
        return miss


# ─── Excel export ─────────────────────────────────────────────────────────────

class ExportService:

    @staticmethod
    def export_to_excel(workspace, start: datetime.date, end: datetime.date) -> bytes:
        """
        Return Excel workbook bytes for the given date range.
        Requires openpyxl.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl is required for Excel export. Add it to pyproject.toml.")

        from apps.crm.models import FinanceEntry, ApplicationEntry, DailySummaryReport

        wb = openpyxl.Workbook()
        ws_sheet = wb.active
        ws_sheet.title = f"{workspace.name} {start}–{end}"

        # Header
        headers = [
            "Дата", "Поступления ($)", "Расходы ($)", "Сальдо ($)",
            "Заработок ПП ($)", "Заработок Привата ($)",
            "Заявки (шт.)", "Заработок с заявок ($)",
            "% плана ПП", "% плана Привата",
        ]
        header_fill = PatternFill(fill_type="solid", fgColor="1E2030")
        header_font = Font(bold=True, color="00D4FF")

        for col, header in enumerate(headers, 1):
            cell = ws_sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        fins  = {e.date: e for e in FinanceEntry.objects.filter(
            workspace=workspace, date__gte=start, date__lte=end
        )}
        apps  = {e.date: e for e in ApplicationEntry.objects.filter(
            workspace=workspace, date__gte=start, date__lte=end
        )}
        reps  = {r.date: r for r in DailySummaryReport.objects.filter(
            workspace=workspace, date__gte=start, date__lte=end
        )}

        current = end
        row = 2
        while current >= start:
            fin = fins.get(current)
            app = apps.get(current)
            rep = reps.get(current)
            ws_sheet.append([
                current.strftime("%d.%m.%Y"),
                float(fin.income)          if fin else "",
                float(fin.expenses)        if fin else "",
                float(fin.balance)         if fin else "",
                float(fin.pp_earnings)     if fin else "",
                float(fin.privat_earnings) if fin else "",
                app.applications_count     if app else "",
                float(app.applications_earnings) if app else "",
                f"{rep.pp_plan_pct:.1f}%"     if rep else "",
                f"{rep.privat_plan_pct:.1f}%"  if rep else "",
            ])
            current -= datetime.timedelta(days=1)
            row += 1

        # Auto-width
        for col in ws_sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_sheet.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

        import io
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
